"""
Motor de dados do CallMed Premium - versao 2, vigente a partir de Setembro/2026.

Regras implementadas (ver Programa_Constelacao_CallMed_v2.md na raiz do projeto para o
documento de referencia completo, aprovado pelo usuario):

- 4 niveis por medico, mensal, por volume de plantoes clinicos validos (exclui Coordenacao/
  Gestao/ASSIST.ADM do volume, exclui hospitais fora do escopo e tipos de unidade fora do
  escopo - ver EXCLUSAO_HOSPITAIS/EXCLUSAO_LOCAL_REGEX).
- Carencia (CORRIGIDO 2026-08-20, esclarecido pelo usuario): meses CONSECUTIVOS cumprindo o
  criterio do nivel so valem pra liberar os BENEFICIOS extras daquele nivel (seguro, reembolso
  de curso, licenca, etc. - ver BENEFICIOS_NIVEL) - carencia de 2 meses pro Nivel 2, 4 meses pro
  Nivel 3, 6 meses pro Nivel 4 (formula: carencia_meses[nivel] + 1 meses seguidos == beneficio
  ativo; cai do nivel em qualquer mes da janela e o contador daquele nivel reseta). O AUMENTO NO
  VALOR DO PLANTAO (pagamento) NAO espera carencia - vale imediatamente com base no volume do
  proprio mes (nivel_bruto). Antes dessa correcao o codigo usava o nivel com carencia
  (nivel_vestido) pra calcular tambem o pagamento, o que segurava o aumento indevidamente (bug
  real encontrado pelo usuario: medico com 22 plantoes/mes, volume mais que suficiente pro Nivel
  4, aparecia preso no Nivel 2 na tela por causa da carencia).
- Coordenadores (qualquer mes com pagamento tipo Coordenacao/Gestao) tem Nivel 4 automatico
  naquele mes, sem precisar de volume nem carencia (pagamento E beneficios, os dois).
- Backfill: o sistema calcula o historico completo desde o inicio da base ate o mes mais
  recente, entao em Setembro/2026 (GO_LIVE) cada medico ja "nasce" no nivel/carencia real que
  o historico dele sustenta - ninguem comeca do zero por definicao de codigo.
"""
import json
import os
import re
from collections import defaultdict
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

import config_caminhos as cfg

GO_LIVE = "2026-09"

# Bump SEMPRE que a LOGICA INTERNA de qualquer funcao de negocio deste modulo mudar (calcular_
# niveis, enriquecer_plantoes, aplicar_operacoes_customizadas, agregar_mensal, etc.) - achado
# real 2026-08-21: as funcoes cacheadas de app.py (@st.cache_data - carregar_linhas_brutas,
# agregar_com_operacoes, calcular_niveis_cached) so invalidam o cache quando o BYTECODE da
# propria funcao DECORADA muda, nao quando uma funcao deste modulo que elas chamam por dentro
# muda - um redeploy do Streamlit Cloud que so atualiza core.py pode deixar o cache antigo vivo,
# servindo resultado ERRADO silenciosamente, sem erro nenhum (bug real: o "colchao" contra queda
# pontual foi pro ar mas o site continuou mostrando o numero de antes por um bom tempo). As 3
# funcoes cacheadas de app.py recebem esse numero como argumento explicito extra justamente pra
# forcar cache miss sempre que ele mudar aqui - bump toda vez que mexer na logica interna.
LOGICA_NEGOCIO_VERSAO = 3  # 3 = expoe meses_abaixo_n2/n3/n4 na saida de calcular_niveis (colunas
# novas, mesmo calculo de antes - ver "risco de queda de beneficio" na pagina Abordagem)

PLACEHOLDERS_MEDICO = {"<Sem Responsável>", "Admin CallmedCall", ".CallMed Adm"}
GESTAO_TIPOS = {"Coordenação", "Gestão", "ASSIST. ADM"}

# Tipos de pagamento que NAO sao plantao clinico novo - bonus/premiacao/adiantamento de um plantao
# que ja tem sua propria linha com Tipo normal em algum lugar. Achado na auditoria de 2026-08-20:
# "Antecipacao" bate com o proprio beneficio documentado do programa ("antecipacao de valores de
# plantao, pago ate 2 dias apos o plantao" - Nivel 1), ou seja, e o MESMO plantao pago cedo, nao
# um plantao a mais. "Callmed Premium" bate com o nome do proprio programa (pagamento do premio de
# fidelidade entrando como linha de "plantao"). Confirmado pelo usuario: excluir da contagem de
# volume/nivel E do valor de repasse (base do % de aumento) - sem isso o medico recebe % de
# aumento em cima do proprio bonus, e o plantao original conta 2x.
# "Bonificacao" e "Gratificacao" ficaram FORA dessa lista (removidas em 2026-08-20, mesmo dia):
# usuario confirmou que sao pagamento por plantao MAIOR que o normal ou por participacao em algo
# atipico da rotina - ou seja, plantao de verdade, so com valor diferenciado. Contam normalmente.
TIPOS_NAO_CONTAM_VOLUME = {
    "Antecipação", "Callmed Premium", "Premiação Infiniti",
    ".Quatro Estrelas CallMed", ".Três Estrelas CallMed", ".Duas Estrelas CallMed",
}

# Hospitais inteiros fora do escopo do programa (qualquer plantao la, de qualquer tipo)
EXCLUSAO_HOSPITAL_REGEX = re.compile(r"covas|amhemed", re.IGNORECASE)
# Escopo do programa = "todas as operacoes de ANESTESIA" (confirmado pelo usuario). A base nova
# (BD Pantoes .../1.ANALISES LUIZ) tem uma coluna "Especialidade" que ja classifica cada linha
# nesse nivel (Anestesia/Enfermaria/UTI/Adm/CallMed/Ambulatorio/Oftalto/...) - usar isso em vez de
# tentar adivinhar UTI/Enfermaria por regex no campo Local (que era fragil, mesmo tipo de
# problema ja visto com a contaminacao do grupo "HSH"). So "Anestesia" conta pro nivel; isso ja
# exclui UTI e Enfermaria automaticamente, sem precisar de regra separada pra elas.
ESPECIALIDADE_VALIDA = "Anestesia"

# Separador da chave composta operacao+especialidade usada na tela "Operacoes" (ex.: "Hospital
# Sao Bernardo :: Anestesia") - precisa ser algo que nao apareca em nome de hospital nem de
# especialidade real da base.
SEP_OP_ESP = " :: "

NIVEIS = [
    # min_fds = minimo de FDS(Sab/Dom) + Noturno somados (uniao) - valores confirmados pelo
    # usuario em 2026-08-20 (exemplo real: medico com 19 plantoes precisa de min. 4).
    # carencia_meses so afeta os BENEFICIOS extras (seguro, curso, licenca - tem_seguro/
    # BENEFICIOS_NIVEL), NAO o pct_aumento do pagamento, que vale imediato por volume
    # (esclarecido pelo usuario 2026-08-20 - ver docstring do modulo). Valores de carencia
    # corrigidos no mesmo dia: Nivel 2 era 1 mes, virou 2; Nivel 3 era 3 meses, virou 4;
    # Nivel 4 continua 6.
    {"idx": 1, "nome": "Nível 1", "min_plantoes": 1, "max_plantoes": 9, "min_fds": 0,
     "carencia_meses": 0, "pct_aumento": 0.0, "pct_exibido": 0, "tem_seguro": False},
    # tem_seguro passou pro Nivel 2 (era so Nivel 3+) - mudanca de regra pedida pelo usuario
    # 2026-08-20. Seguro agora entra no programa a partir daqui.
    {"idx": 2, "nome": "Nível 2", "min_plantoes": 10, "max_plantoes": 14, "min_fds": 3,
     "carencia_meses": 2, "pct_aumento": 0.0275, "pct_exibido": 3, "tem_seguro": True},
    {"idx": 3, "nome": "Nível 3", "min_plantoes": 15, "max_plantoes": 19, "min_fds": 4,
     "carencia_meses": 4, "pct_aumento": 0.0375, "pct_exibido": 4, "tem_seguro": True},
    {"idx": 4, "nome": "Nível 4", "min_plantoes": 20, "max_plantoes": None, "min_fds": 5,
     "carencia_meses": 6, "pct_aumento": 0.045, "pct_exibido": 5, "tem_seguro": True},
]
NIVEL_POR_IDX = {n["idx"]: n for n in NIVEIS}

# Custos reais confirmados por apolice (Porto Seguro + Unimed Seguros), por medico/mes, a partir
# do Nivel 2 (onde o seguro entra no programa - mudou de Nivel 3 pro Nivel 2, pedido do usuario
# 2026-08-20).
CUSTO_SEGURO_VIDA_DIT_FUNERAL = 185.61  # Porto Seguro: Vida (99k/198k) + DIT (166,66/dia) + Funeral (10k)
CUSTO_RCP = 117.60  # Unimed Seguros: RCP 100k, franquia zero
CUSTO_SEGURO_TOTAL_MES = CUSTO_SEGURO_VIDA_DIT_FUNERAL + CUSTO_RCP  # R$303,21


def _nivel_bruto(n_plantoes, n_fds, niveis=None):
    """Nivel que o volume/fds do mes sustenta, sem considerar carencia nem coordenacao."""
    niveis = niveis or NIVEIS
    nivel = 1
    for n in niveis:
        maximo = n["max_plantoes"] if n["max_plantoes"] is not None else float("inf")
        if n["min_plantoes"] <= n_plantoes <= maximo and n_fds >= n["min_fds"]:
            nivel = n["idx"]
    return nivel


def niveis_para_dict(niveis=None):
    return {n["idx"]: n for n in (niveis or NIVEIS)}


def carregar_apoio(arquivo=None):
    """Le a aba 'Apoio' do arquivo de plantoes - tabela de referencia mantida pela area que
    mapeia cada Local EXATO pro Coordenador responsavel, o 'Setor Definido' (agrupamento correto
    de hospital/operacao) e a Especialidade correta.

    Achado na auditoria de 2026-08-20 (usuario pediu pra cruzar contra a propria base de
    plantoes): a coluna 'Especialidade' da aba BD tem erro de classificacao linha a linha - ex.:
    o mesmo Local exato, no mesmo dia, aparece ora como Enfermaria ora como Anestesia. A aba Apoio
    e a fonte de verdade oficial (mantida como lookup Local -> Setor/Especialidade), nao a coluna
    Especialidade da BD - confirmado: BD diverge da Apoio em 4.102 das 181.423 linhas (~2,3%), nos
    dois sentidos (tanto rotula Enfermaria/UTI como Anestesia quanto o contrario). Cobertura
    tambem confirmada: 100% dos Local distintos que aparecem na BD tem entrada aqui.

    Bonus: 'Setor Definido' tambem resolve a fragmentacao de nomes que operacao_de() (regex sobre
    o Local) nao pegava - ex.: 'Ana Costa, Hospital', 'Hospital Ana Costa' e 'Hospital Ana Costa -
    P2' (o hospital foi renomeado no cadastro em 2025-11, mesma rede) todos caem no mesmo Setor
    Definido 'Ana Costa Anestesia'."""
    arquivo = arquivo or cfg.resolver_base_plantoes()
    wb = load_workbook(arquivo, data_only=True, read_only=True)
    if "Apoio" not in wb.sheetnames:
        wb.close()
        return pd.DataFrame(columns=["local", "coordenador", "setor_definido", "especialidade_apoio"])
    ws = wb["Apoio"]
    rows_iter = ws.iter_rows(values_only=True)
    next(rows_iter)  # cabecalho
    linhas = []
    for row in rows_iter:
        local = row[0] if len(row) > 0 else None
        if local is None:
            continue
        linhas.append({
            "local": str(local),
            "coordenador": str(row[1]) if len(row) > 1 and row[1] is not None else "",
            "setor_definido": str(row[2]) if len(row) > 2 and row[2] is not None else "",
            "especialidade_apoio": str(row[3]) if len(row) > 3 and row[3] is not None else "",
        })
    wb.close()
    apoio = pd.DataFrame(linhas)
    if apoio.empty:
        return apoio
    return apoio.drop_duplicates("local", keep="first")


_COLUNAS_APOIO = ["local", "coordenador", "setor_definido", "especialidade_apoio"]


def salvar_apoio_customizado(apoio_df, caminho=None):
    """Persiste o mapeamento Local -> Setor Definido/Especialidade GERENCIADO PELO SISTEMA (tela
    '🗂️ Apoio', 2026-08-20) em disco como JSON - sobrevive a reruns/reinicios locais do servidor
    (ver config_caminhos.APOIO_CUSTOMIZADO e a ressalva sobre Streamlit Cloud la)."""
    caminho = caminho or cfg.APOIO_CUSTOMIZADO
    os.makedirs(os.path.dirname(caminho), exist_ok=True)
    registros = apoio_df[_COLUNAS_APOIO].to_dict("records")
    with open(caminho, "w", encoding="utf-8") as f:
        json.dump(registros, f, ensure_ascii=False, indent=2)


def carregar_apoio_customizado(caminho=None):
    """Le o mapeamento Local -> Setor/Especialidade persistido pela tela '🗂️ Apoio' (JSON), se
    ja existir. DataFrame vazio se ainda nao foi criado (antes do primeiro bootstrap)."""
    caminho = caminho or cfg.APOIO_CUSTOMIZADO
    if not os.path.exists(caminho):
        return pd.DataFrame(columns=_COLUNAS_APOIO)
    with open(caminho, "r", encoding="utf-8") as f:
        registros = json.load(f)
    apoio = pd.DataFrame(registros, columns=_COLUNAS_APOIO)
    if apoio.empty:
        return pd.DataFrame(columns=_COLUNAS_APOIO)
    return apoio.drop_duplicates("local", keep="first")


def resolver_apoio(arquivo=None):
    """De onde vem o mapeamento Local -> Setor/Especialidade usado no calculo: se ja existe uma
    versao GERENCIADA PELO SISTEMA (persistida pela tela '🗂️ Apoio'), usa ela - senao importa da
    aba 'Apoio' do Excel uma unica vez (bootstrap) e ja salva, pro sistema assumir dai em diante
    sem depender mais da planilha. Pedido do usuario em 2026-08-20: 'ao inves de consultar a
    planilha, essa regra e gerida pelo sistema'."""
    apoio_sistema = carregar_apoio_customizado()
    if not apoio_sistema.empty:
        return apoio_sistema
    apoio_planilha = carregar_apoio(arquivo)
    if not apoio_planilha.empty:
        salvar_apoio_customizado(apoio_planilha)
    return apoio_planilha


def resumo_locais_para_apoio(df_linhas, apoio_df):
    """Junta os Local distintos da base (com volume de linhas e nº de médicos) com o mapeamento
    atual (sistema ou planilha) - usado pra montar a tela '🗂️ Apoio', pro master ver o impacto de
    cada Local antes de editar e priorizar o que falta classificar (sobe pro topo, ordenado por
    volume dentro de cada grupo)."""
    colunas_vazias = ["local", "total_linhas", "medicos", "coordenador", "setor_definido",
                       "especialidade_apoio", "classificado"]
    if df_linhas.empty:
        return pd.DataFrame(columns=colunas_vazias)
    resumo = df_linhas.groupby("local").agg(
        total_linhas=("valor", "count"), medicos=("medico", "nunique"),
    ).reset_index()
    apoio_df = apoio_df if apoio_df is not None and not apoio_df.empty else pd.DataFrame(columns=_COLUNAS_APOIO)
    tabela = resumo.merge(apoio_df, on="local", how="left")
    for col in ("coordenador", "setor_definido", "especialidade_apoio"):
        tabela[col] = tabela[col].fillna("")
    tabela["classificado"] = tabela["setor_definido"] != ""
    return tabela.sort_values(["classificado", "total_linhas"], ascending=[True, False])


class ArquivoInvalidoError(Exception):
    """Levantado quando o arquivo enviado (upload via '📁 Fonte de dados') não tem o formato
    esperado - aba 'BD' ausente, planilha vazia, etc. Achado real 2026-08-20: o master enviou uma
    planilha sem essa aba e o sistema quebrou com uma tela de traceback em vez de um aviso
    entendível - core.py levanta isso com uma mensagem clara, app.py mostra via st.error()."""


def _ler_plantoes_excel(arquivo=None, retornar_stats=False):
    """Le a planilha crua e devolve um DataFrame com as colunas base (anomes/medico/data_raw/
    local/tipo/especialidade_bd/valor), sem nenhum flag de elegibilidade calculado ainda - ver
    enriquecer_plantoes() pro resto do pipeline. Detecta automaticamente qual dos 2 formatos
    aceitos a planilha usa, pela aba presente:

    - Aba 'BD' (formato "1.ANALISES LUIZ", histórico/legado - fallback offline, migração inicial).
    - Aba 'Plantões' (export do sistema de escala "pegaplantao.com.br" - formato usado a partir de
      2026-08-20 pra todo upload feito pela tela '📁 Fonte de dados', pedido do usuário: "segue o
      formato da planilha que vamos sempre mandar"). Não tem colunas mes/ano separadas (deriva de
      'Data') nem 'Especialidade' (especialidade_bd fica vazia - sem problema, a aba Apoio sempre
      tem prioridade sobre ela no cálculo real, ver enriquecer_plantoes()). A coluna 'Profissional
      de Plantão' é a fonte de verdade de quem realmente cobriu o plantão - usada como 'medico'
      independente do que a coluna 'Situação' diz (ex.: 'Trocado'/'Troca de Fixo'/'Falta
      Justificada' também usam essa coluna, sem filtro extra - confirmado pelo usuário 2026-08-20:
      "considere a coluna profissional de plantão a real... tem que ser essa a referência pra
      tudo"). 'Situação' em si não é lida.

    Se nenhuma das duas abas existir, levanta ArquivoInvalidoError com uma mensagem clara (lista
    as abas encontradas) em vez de deixar o KeyError cru do openpyxl derrubar a tela (achado real
    2026-08-20: upload de planilha em formato errado quebrava o app inteiro com um traceback)."""
    arquivo = arquivo or cfg.resolver_base_plantoes()
    wb = load_workbook(arquivo, data_only=True, read_only=True)
    if "BD" in wb.sheetnames:
        df, stats = _ler_plantoes_excel_formato_bd(wb)
    elif "Plantões" in wb.sheetnames:
        df, stats = _ler_plantoes_excel_formato_pegaplantao(wb)
    else:
        abas_encontradas = ", ".join(wb.sheetnames) if wb.sheetnames else "nenhuma"
        wb.close()
        raise ArquivoInvalidoError(
            f"Essa planilha não tem uma aba 'BD' nem 'Plantões' (abas encontradas: "
            f"{abas_encontradas}). Confira se é o arquivo certo."
        )
    if retornar_stats:
        return df, stats
    return df


def _ler_plantoes_excel_formato_bd(wb):
    """Aba 'BD' (formato "1.ANALISES LUIZ") - histórico/legado, mantido como fallback offline.
    Linhas com valor <= 0 (ou nao numerico) sao descartadas - normalmente lixo/linha em branco,
    mas pode incluir uma correcao negativa legitima tipo 'Diferença Vr de plantão' (ainda em
    aberto, ver Programa_Constelacao_CallMed_v2.md secao 4)."""
    ws = wb["BD"]
    rows_iter = ws.iter_rows(values_only=True)
    next(rows_iter)
    linhas = []
    descartadas_valor_invalido = 0
    for row in rows_iter:
        mes, ano, data_raw, local = row[3], row[4], row[5], row[7]
        medico, tipo, valor = row[10], row[11], row[12]
        especialidade = row[18] if len(row) > 18 else None
        if not isinstance(mes, (int, float)) or not isinstance(ano, (int, float)):
            continue
        if not isinstance(valor, (int, float)) or valor <= 0:
            if medico and medico not in PLACEHOLDERS_MEDICO:
                # so conta como "descartada" se pelo menos parecia uma linha de verdade (tinha
                # medico) - uma linha em branco de fato nao e um dado perdido, e ruido da planilha.
                descartadas_valor_invalido += 1
            continue
        if not medico or medico in PLACEHOLDERS_MEDICO:
            continue
        local_s = str(local or "")
        linhas.append({
            "anomes": f"{int(ano)}-{int(mes):02d}",
            "medico": str(medico).strip(),
            "data_raw": data_raw,
            "local": local_s,
            "tipo": str(tipo or ""),
            "especialidade_bd": str(especialidade or ""),
            "valor": float(valor),
        })
    wb.close()
    df = pd.DataFrame(linhas)
    return df, {"descartadas_valor_invalido": descartadas_valor_invalido}


_COLUNAS_OBRIGATORIAS_PEGAPLANTAO = ["Data", "Local", "Profissional de Plantão", "Tipo", "Valor"]


def _ler_plantoes_excel_formato_pegaplantao(wb):
    """Aba 'Plantões' - export do sistema de escala "pegaplantao.com.br", formato padrão a partir
    de 2026-08-20. Tem algumas linhas de metadado antes do cabeçalho de verdade (título, período,
    'gerado em...') - a quantidade pode variar entre exports, então acha o cabeçalho procurando a
    linha cuja 1ª célula é literalmente 'Data', em vez de pular um número fixo de linhas. Também
    tem uma linha de rodapé 'Total Geral' (sem médico - já cai fora pelo filtro de médico vazio) e
    um aviso de autenticidade no fim (mesma coisa, sem médico)."""
    ws = wb["Plantões"]
    rows_iter = ws.iter_rows(values_only=True)

    header = None
    for row in rows_iter:
        primeira = str(row[0]).strip() if row and row[0] is not None else ""
        if primeira == "Data":
            header = row
            break
    if header is None:
        wb.close()
        raise ArquivoInvalidoError(
            "Não encontrei a linha de cabeçalho (coluna 'Data') na aba 'Plantões' dessa planilha. "
            "Confira se é o arquivo certo, exportado do sistema de escala."
        )
    col_idx = {str(h).strip(): i for i, h in enumerate(header) if h is not None}
    faltando = [c for c in _COLUNAS_OBRIGATORIAS_PEGAPLANTAO if c not in col_idx]
    if faltando:
        wb.close()
        raise ArquivoInvalidoError(
            f"A aba 'Plantões' dessa planilha não tem a(s) coluna(s) esperada(s): "
            f"{', '.join(faltando)}. Confira se é o arquivo certo."
        )
    i_data = col_idx["Data"]
    i_local = col_idx["Local"]
    i_medico = col_idx["Profissional de Plantão"]
    i_tipo = col_idx["Tipo"]
    i_valor = col_idx["Valor"]
    n_colunas_min = max(i_data, i_local, i_medico, i_tipo, i_valor) + 1

    linhas = []
    descartadas_valor_invalido = 0
    for row in rows_iter:
        if len(row) < n_colunas_min:
            continue
        data_raw, local, medico, tipo, valor = (
            row[i_data], row[i_local], row[i_medico], row[i_tipo], row[i_valor]
        )
        if not isinstance(valor, (int, float)) or valor <= 0:
            if medico and medico not in PLACEHOLDERS_MEDICO:
                descartadas_valor_invalido += 1
            continue
        if not medico or medico in PLACEHOLDERS_MEDICO:
            continue
        # Sem colunas mes/ano separadas nesse formato - deriva da propria Data (mesmo texto
        # "dd/mm/aaaa hh:mm" usado como data_raw, dayfirst=True igual ao resto do sistema).
        data_dt = pd.to_datetime(data_raw, dayfirst=True, errors="coerce")
        if pd.isna(data_dt):
            continue
        linhas.append({
            "anomes": f"{data_dt.year}-{data_dt.month:02d}",
            "medico": str(medico).strip(),
            "data_raw": data_raw,
            "local": str(local or ""),
            "tipo": str(tipo or ""),
            "especialidade_bd": "",
            "valor": float(valor),
        })
    wb.close()
    df = pd.DataFrame(linhas)
    return df, {"descartadas_valor_invalido": descartadas_valor_invalido}


def enriquecer_plantoes(df, apoio_df=None, arquivo=None):
    """Recebe o DataFrame CRU (colunas anomes/medico/data_raw/local/tipo/especialidade_bd/valor -
    de _ler_plantoes_excel() OU de consultar_plantoes_supabase()) e calcula todos os flags de
    elegibilidade (FDS/noturno, gestao, tipo nao-clinico, operacao/especialidade via Apoio,
    conta_pro_nivel). Compartilhado pelas duas fontes de dados, pra nao duplicar a regra."""
    if df.empty:
        return df
    df = df.copy()
    df["data_dt"] = pd.to_datetime(df["data_raw"], dayfirst=True, errors="coerce")
    # FDS = Sabado e Domingo (sexta NAO conta mais - mudanca de regra confirmada pelo usuario,
    # sessao 2026-08-20, ver Programa_Constelacao_CallMed_v2.md).
    df["eh_fds"] = df["data_dt"].dt.dayofweek.isin([5, 6]).fillna(False)
    # Noturno = comeca as 19h ou depois, OU esta marcado como Noturno/Cinderela no Tipo/Local
    # (mesma sessao) - cobre casos com esse rotulo mas horario de inicio um pouco antes das 19h.
    texto_tipo_local = (df["tipo"].fillna("") + " " + df["local"].fillna("")).str.lower()
    df["eh_noturno"] = (
        (df["data_dt"].dt.hour >= 19).fillna(False)
        | texto_tipo_local.str.contains("noturno|cinderela", regex=True, na=False)
    )
    # Exigencia de nivel (min_fds) passa a valer sobre a UNIAO de FDS e Noturno, sem contar 2x um
    # plantao que seja os dois (ex.: sexta a noite - sexta ja nao conta como FDS pela regra nova,
    # mas se fosse sabado a noite contaria 1 vez so aqui).
    df["eh_fds_ou_noturno"] = df["eh_fds"] | df["eh_noturno"]
    df["eh_gestao"] = df["tipo"].isin(GESTAO_TIPOS)
    df["eh_tipo_nao_clinico"] = df["tipo"].isin(TIPOS_NAO_CONTAM_VOLUME)
    df["operacao_bd"] = df["local"].apply(operacao_de)

    # Junta com o mapeamento Local -> Setor Definido/Especialidade - fonte de verdade pro
    # "operacao"/"especialidade" corrigidos ("operacao_bd"/"especialidade_bd" ficam guardadas so
    # pra auditoria/transparencia, nunca usadas no calculo de nivel). 'apoio_df' permite passar
    # a versao GERENCIADA PELO SISTEMA (tela '🗂️ Apoio', editavel pelo master) em vez de sempre
    # reler a aba Apoio do Excel - se nao vier nada, resolver_apoio() decide (sistema, se ja
    # existir, senao a planilha como bootstrap).
    apoio = apoio_df if apoio_df is not None else resolver_apoio(arquivo)
    if not apoio.empty:
        df = df.merge(apoio, on="local", how="left")
    else:
        df["setor_definido"] = None
        df["especialidade_apoio"] = None
        df["coordenador"] = None
    # Fallback conservador pra Local sem entrada na Apoio (cobertura hoje e 100%, mas protege
    # contra um Local novo aparecer numa atualizacao futura da BD sem a Apoio ter acompanhado) -
    # cai pro agrupamento por regex + Especialidade da propria BD, e fica marcado pra tela avisar.
    df["sem_match_apoio"] = df["setor_definido"].isna() | (df["setor_definido"] == "")
    df["operacao"] = df["setor_definido"].where(~df["sem_match_apoio"], df["operacao_bd"])
    df["especialidade"] = df["especialidade_apoio"].where(~df["sem_match_apoio"], df["especialidade_bd"])

    df["hospital_excluido"] = df["local"].str.contains(EXCLUSAO_HOSPITAL_REGEX, na=False)
    df["eh_anestesia"] = df["especialidade"] == ESPECIALIDADE_VALIDA
    # "chave_operacao" = operacao (Setor Definido) + especialidade (Apoio), ex.: "Ana Costa
    # Anestesia :: Anestesia" - granularidade real da tela "Operacoes".
    especialidade_disp = df["especialidade"].replace("", "(sem especialidade)")
    df["chave_operacao"] = df["operacao"] + SEP_OP_ESP + especialidade_disp
    # plantao valido = conta pro volume do nivel: especialidade Anestesia (Apoio), hospital nao
    # excluido, e nao e pagamento de gestao/coordenacao nem tipo nao-clinico. Este e o padrao
    # default - a tela "Operacoes" permite substituir esse filtro por uma lista customizada de
    # chave_operacao via aplicar_operacoes_customizadas(), sem precisar reler o Excel.
    df["conta_pro_nivel"] = (
        df["eh_anestesia"] & (~df["hospital_excluido"]) & (~df["eh_gestao"])
        & (~df["eh_tipo_nao_clinico"])
    )
    return df


def carregar_plantoes(arquivo=None, apoio_df=None):
    """Le a base de plantoes do EXCEL local (aba 'BD') e ja enriquece. Mantida como caminho de
    fallback/offline (scripts, migracao) - a tela do sistema usa consultar_plantoes_supabase()
    desde 2026-08-20 (le do Supabase, nao depende mais de acesso a maquina/OneDrive)."""
    df_cru = _ler_plantoes_excel(arquivo)
    return enriquecer_plantoes(df_cru, apoio_df=apoio_df, arquivo=arquivo)


_PAGINA_SUPABASE = 1000


def consultar_plantoes_supabase(client, apoio_df=None):
    """Le a base de plantoes inteira do Supabase (tabela public.plantoes, paginada de 1000 em
    1000 - limite padrao do PostgREST) e ja enriquece. Fonte PRINCIPAL desde 2026-08-20 - o
    sistema para de depender do Excel local/OneDrive pra rodar (pedido do usuario: "não dependo
    de consulta na máquina"). 'client' e o objeto supabase-py (ver supabase_client.get_client())."""
    linhas = []
    offset = 0
    while True:
        resposta = (
            client.table("plantoes")
            .select("anomes,medico,data_raw,local,tipo,especialidade_bd,valor")
            .order("id")  # obrigatorio pra paginacao estavel - sem ORDER BY o Postgres nao
            # garante ordem consistente entre paginas, o que pode pular ou duplicar linha
            .range(offset, offset + _PAGINA_SUPABASE - 1)
            .execute()
        )
        pagina = resposta.data
        if not pagina:
            break
        linhas.extend(pagina)
        if len(pagina) < _PAGINA_SUPABASE:
            break
        offset += _PAGINA_SUPABASE
    df_cru = pd.DataFrame(linhas)
    if not df_cru.empty:
        # data_raw volta do Supabase ja em ISO (yyyy-mm-ddThh:mm:ss), inequivoco - NAO pode
        # passar por dayfirst=True de novo dentro de enriquecer_plantoes() (que existe pro
        # formato textual "dd/mm/aaaa hh:mm" do Excel), senao mes/dia trocam de lugar quando os
        # dois sao <=12 (ex.: 2023-04-01 vira 2023-01-04) ou a data falha o parse inteira (bug
        # real encontrado e corrigido em 2026-08-20). Convertendo aqui pra Timestamp de verdade
        # ANTES, o pd.to_datetime(dayfirst=True) la dentro vira passthrough (input ja e
        # datetime64, dayfirst e ignorado nesse caso).
        df_cru["data_raw"] = pd.to_datetime(df_cru["data_raw"], errors="coerce")
    return enriquecer_plantoes(df_cru, apoio_df=apoio_df)


def consultar_apoio_supabase(client):
    """Le o mapeamento Local -> Setor/Especialidade da tabela public.apoio do Supabase - versao
    GERENCIADA PELO SISTEMA (substitui carregar_apoio_customizado()/resolver_apoio() locais desde
    a migracao pro Supabase, 2026-08-20)."""
    resposta = client.table("apoio").select("local,coordenador,setor_definido,especialidade").execute()
    apoio = pd.DataFrame(resposta.data)
    if apoio.empty:
        return pd.DataFrame(columns=_COLUNAS_APOIO)
    return apoio.rename(columns={"especialidade": "especialidade_apoio"})[_COLUNAS_APOIO]


def salvar_apoio_supabase(client, apoio_df):
    """Grava o mapeamento Local -> Setor/Especialidade editado na tela '🗂️ Apoio' de volta no
    Supabase (upsert por 'local', a chave primaria da tabela)."""
    registros = apoio_df.rename(columns={"especialidade_apoio": "especialidade"})[
        ["local", "coordenador", "setor_definido", "especialidade"]
    ].to_dict("records")
    for i in range(0, len(registros), 500):
        client.table("apoio").upsert(registros[i:i + 500], on_conflict="local").execute()


def consultar_config_supabase(client, chave, padrao=None):
    """Le uma config editavel (niveis_custom, operacoes_excluidas, medicos_gestores, custo_seguro,
    rampup) da tabela public.config (JSONB, chave->valor) - substitui o que antes vivia só em
    st.session_state e se perdia a cada sessão nova/redeploy (achado real na auditoria de
    2026-08-20: 5 telas de configuração "aplicavam" mas não gravavam em lugar nenhum permanente).
    Retorna 'padrao' se a chave ainda não existe (primeira vez que o sistema roda)."""
    resposta = client.table("config").select("valor").eq("chave", chave).limit(1).execute()
    if not resposta.data:
        return padrao
    return resposta.data[0]["valor"]


def salvar_config_supabase(client, chave, valor, alterado_por=None):
    """Grava (upsert) uma config editável na tabela public.config, e registra a mudança em
    public.config_historico (quem, quando, valor antes/depois) - pedido do usuário 2026-08-21,
    pra dar pra reconstruir "quais eram as regras vigentes num dia X" mesmo depois de editadas
    (sem isso, editar uma regra reescreve o histórico de pagamento inteiro sem deixar rastro).
    'alterado_por' é o nome/e-mail de quem está logado - None se não vier (ex.: chamadas de
    script/migração)."""
    valor_anterior = consultar_config_supabase(client, chave, None)
    client.table("config_historico").insert({
        "chave": chave, "valor_anterior": valor_anterior, "valor_novo": valor,
        "alterado_por": alterado_por,
    }).execute()
    client.table("config").upsert(
        {"chave": chave, "valor": valor}, on_conflict="chave"
    ).execute()


def consultar_config_historico_supabase(client, limite=100):
    """Le o log de auditoria das mudanças de configuração (tabela public.config_historico),
    mais recente primeiro - usado na tela '⚙️ Regras do Programa' pra transparência de quem
    mudou o quê e quando."""
    resposta = (
        client.table("config_historico")
        .select("chave,valor_anterior,valor_novo,alterado_por,alterado_em")
        .order("alterado_em", desc=True)
        .limit(limite)
        .execute()
    )
    return pd.DataFrame(
        resposta.data,
        columns=["chave", "valor_anterior", "valor_novo", "alterado_por", "alterado_em"],
    )


def consultar_rampup_supabase(client):
    """Le todos os disparos de bônus de ramp-up já confirmados (tabela public.rampup_disparos) -
    cada linha é um disparo real, com a operação, o % e a lista exata de meses escolhidos."""
    resposta = (
        client.table("rampup_disparos")
        .select("id,operacao,pct,meses,disparado_por,disparado_em")
        .order("disparado_em", desc=True)
        .execute()
    )
    return pd.DataFrame(
        resposta.data,
        columns=["id", "operacao", "pct", "meses", "disparado_por", "disparado_em"],
    )


def salvar_rampup_supabase(client, operacao, pct, meses, disparado_por):
    """Grava um disparo de bônus de ramp-up de verdade (antes só mostrava uma mensagem de sucesso
    sem persistir nada em lugar nenhum - achado real na auditoria de 2026-08-20, o botão
    'Confirmar disparo' não tinha efeito nenhum além da mensagem na tela)."""
    client.table("rampup_disparos").insert({
        "operacao": operacao, "pct": float(pct), "meses": list(meses),
        "disparado_por": disparado_por,
    }).execute()


def calcular_rampup_por_medico_mes(df_linhas, rampup_df):
    """Para cada disparo de ramp-up (operação + lista de meses + %), soma o valor de repasse do
    médico NAQUELA operação específica, só nos meses escolhidos no disparo, e aplica o %. Um
    médico que atende mais de um hospital em ramp-up ao mesmo tempo soma os bônus de cada um.
    Retorna (medico, anomes) -> custo_rampup_mes, pra o app somar em cima de valor_total_geral."""
    colunas_vazias = ["medico", "anomes", "custo_rampup_mes"]
    if rampup_df is None or rampup_df.empty or df_linhas.empty:
        return pd.DataFrame(columns=colunas_vazias)
    validos = df_linhas[df_linhas["conta_pro_nivel"]]
    partes = []
    for _, disparo in rampup_df.iterrows():
        meses_disparo = set(disparo["meses"] or [])
        if not meses_disparo:
            continue
        filtro = (validos["operacao"] == disparo["operacao"]) & validos["anomes"].isin(meses_disparo)
        sub = validos[filtro].groupby(["medico", "anomes"], as_index=False)["valor"].sum()
        sub["custo_rampup_mes"] = sub["valor"] * float(disparo["pct"])
        partes.append(sub[["medico", "anomes", "custo_rampup_mes"]])
    if not partes:
        return pd.DataFrame(columns=colunas_vazias)
    todos = pd.concat(partes, ignore_index=True)
    return todos.groupby(["medico", "anomes"], as_index=False)["custo_rampup_mes"].sum()


def tipos_novos_no_mes(df_linhas, mes_ref):
    """Tipos de pagamento (coluna 'Tipo') vistos no mes de referencia que NUNCA apareceram em
    nenhum mes anterior da base inteira - sinal de que o sistema de escala (pegaplantao.com.br)
    pode ter renomeado ou criado um Tipo novo, que ainda nao foi revisado pra saber se deve
    contar ou nao pro nivel (TIPOS_NAO_CONTAM_VOLUME/GESTAO_TIPOS sao comparacao exata de string -
    um Tipo renomeado silenciosamente passaria a contar errado, sem crash nenhum pra avisar).
    Pedido do usuario 2026-08-21: pegar isso na hora, nao 3 meses depois."""
    if df_linhas.empty:
        return []
    tipos_mes = set(df_linhas.loc[df_linhas["anomes"] == mes_ref, "tipo"].unique())
    tipos_antes = set(df_linhas.loc[df_linhas["anomes"] < mes_ref, "tipo"].unique())
    return sorted(t for t in (tipos_mes - tipos_antes) if t)


def enviar_planilha_supabase(client, arquivo_upload):
    """Le uma planilha (upload do master, cobrindo qualquer periodo - nao precisa ser a base
    inteira) e faz UPSERT no Supabase, usando a chave (medico, data_raw, local, tipo, valor) pra
    decidir se cada linha e nova ou ja existia (mesma linha de novo = sobreposicao, so confirma o
    que ja tinha; combinacao nova = incluida). Pedido do usuario: "eu mando os dados via upload de
    uma planilha de um período que eu desejar e você avalia se tem sobreposição ou se inclui dados
    novos".

    Retorna um resumo (dict) pra tela mostrar pro master: quantas linhas vieram na planilha, quais
    meses ela cobre, quantas ja existiam no banco pra esses meses (antes do envio) e quantas
    existem depois - a diferenca e quanto foi realmente adicionado."""
    df_cru, stats_leitura = _ler_plantoes_excel(arquivo_upload, retornar_stats=True)
    if df_cru.empty:
        return {
            "linhas_enviadas": 0, "meses": [], "existentes_antes": 0, "existentes_depois": 0,
            "novas": 0, "descartadas_valor_invalido": stats_leitura["descartadas_valor_invalido"],
        }

    meses = sorted(df_cru["anomes"].unique())

    def _contar_existentes(meses):
        total = 0
        for mes in meses:
            resp = (
                client.table("plantoes").select("id", count="exact")
                .eq("anomes", mes).limit(1).execute()
            )
            total += resp.count or 0
        return total

    existentes_antes = _contar_existentes(meses)

    registros = df_cru.copy()
    # "mes"/"ano" nao vem de _ler_plantoes_excel() (que so tem "anomes" combinado) mas a tabela
    # exige os dois (not null) - deriva de volta a partir de anomes ("aaaa-mm").
    registros["ano"] = registros["anomes"].str[:4].astype(int)
    registros["mes"] = registros["anomes"].str[5:7].astype(int)
    # data_raw normalmente vem como TEXTO ("dd/mm/aaaa hh:mm"), nao datetime nativo - mesmo
    # parsing de enriquecer_plantoes() (dayfirst=True), senao a data se perde (bug ja visto e
    # corrigido na migracao inicial de 2026-08-20).
    registros["data_raw"] = pd.to_datetime(
        registros["data_raw"], dayfirst=True, errors="coerce"
    ).apply(lambda d: d.isoformat() if pd.notna(d) else None)
    # Duplicata exata nos 5 campos da chave DENTRO do proprio upload derruba o upsert (Postgres
    # nao deixa um UPSERT afetar a mesma linha 2x no mesmo comando) - risco aceito pelo usuario ao
    # escolher essa chave, mantem so a 1a ocorrencia.
    registros = registros.drop_duplicates(
        subset=["medico", "data_raw", "local", "tipo", "valor"], keep="first"
    )
    registros = registros.to_dict("records")
    for i in range(0, len(registros), 2000):
        lote = registros[i:i + 2000]
        client.table("plantoes").upsert(
            lote, on_conflict="medico,data_raw,local,tipo,valor"
        ).execute()

    existentes_depois = _contar_existentes(meses)

    return {
        "linhas_enviadas": len(df_cru),
        "meses": meses,
        "existentes_antes": existentes_antes,
        "existentes_depois": existentes_depois,
        "descartadas_valor_invalido": stats_leitura["descartadas_valor_invalido"],
        "novas": existentes_depois - existentes_antes,
    }


def operacao_de(local):
    """Extrai o nome da 'operacao' (hospital/grupo) de um campo Local - mesmo criterio usado no
    seletor de ramp-up: remove so o ultimo segmento (turno/sub-unidade), ex.: 'Hospital Ana Costa
    - Centro Cirurgico' -> 'Hospital Ana Costa'."""
    return str(local or "").rsplit(" - ", 1)[0].strip()


def listar_operacoes(df):
    """Lista as combinacoes distintas de (operacao, especialidade) encontradas na base - operacao
    e especialidade ja sao as versoes corrigidas pela aba Apoio (ver carregar_apoio()), nao a
    coluna Especialidade crua da BD nem o agrupamento por regex. Cada combinacao e uma linha
    flegavel separada na tela de Configuracoes > Operacoes."""
    colunas_vazias = ["chave_operacao", "operacao", "especialidade", "total_linhas", "conta_hoje",
                       "incluida_padrao"]
    if df.empty:
        return pd.DataFrame(columns=colunas_vazias)
    resumo = df.groupby(["chave_operacao", "operacao", "especialidade"]).agg(
        total_linhas=("valor", "count"),
        conta_hoje=("conta_pro_nivel", "sum"),
    ).reset_index()
    resumo["especialidade"] = resumo["especialidade"].replace("", "(sem especialidade)")
    resumo["incluida_padrao"] = resumo["conta_hoje"] > 0
    return resumo.sort_values(["operacao", "especialidade"])


def operacoes_excluidas_por_padrao(df):
    """Calcula o conjunto de chaves (operacao+especialidade) que o filtro padrao excluiria hoje:
    so Anestesia conta, e dentro disso ainda exclui os hospitais do EXCLUSAO_HOSPITAL_REGEX -
    usado so pra inicializar a tela de Configuracoes > Operacoes com um estado EQUIVALENTE ao
    comportamento anterior ao refinamento por especialidade, antes do master customizar."""
    if df.empty:
        return set()
    elegivel_padrao = df["eh_anestesia"] & (~df["hospital_excluido"])
    return set(df.loc[~elegivel_padrao, "chave_operacao"].unique())


def aplicar_operacoes_customizadas(df, operacoes_excluidas):
    """Recalcula conta_pro_nivel usando um conjunto customizado de chaves (operacao+especialidade)
    excluidas (tela de Configuracoes > Operacoes) em vez do filtro fixo (so Anestesia, fora
    EXCLUSAO_HOSPITAL_REGEX). Granularidade agora e por combinacao especifica, entao o master pode
    liberar uma especialidade fora de Anestesia se quiser (ela so nao aparece pre-marcada por
    padrao). Sempre exclui eh_gestao (coordenacao/gestao) e eh_tipo_nao_clinico (bonus/premiacao/
    adiantamento - ver TIPOS_NAO_CONTAM_VOLUME) independente da tela de Operacoes, porque esses
    dois nao sao sobre ESCOPO de hospital/especialidade e sim sobre o que conta como plantao de
    verdade. Nao mexe no Excel, so reprocessa o dataframe ja carregado."""
    df = df.copy()
    operacoes_excluidas = set(operacoes_excluidas or [])
    df["chave_excluida"] = df["chave_operacao"].isin(operacoes_excluidas)
    df["conta_pro_nivel"] = (
        (~df["chave_excluida"]) & (~df["eh_gestao"]) & (~df["eh_tipo_nao_clinico"])
    )
    return df


def agregar_mensal(df):
    """Agrega linha-a-linha em (medico, anomes) -> n_plantoes validos, n_fds (Sab/Dom) validos,
    n_noturno validos, n_fds_ou_noturno (uniao, base da exigencia de nivel), se teve pagamento de
    coordenacao/gestao naquele mes (mesmo fora do escopo de hospital - coordenacao e coordenacao
    em qualquer lugar), e valor de repasse elegivel (base pro % de aumento)."""
    colunas_vazias = ["medico", "anomes", "n_plantoes", "n_fds", "n_noturno", "n_fds_ou_noturno",
                       "teve_coordenacao", "valor_repasse"]
    if df.empty:
        return pd.DataFrame(columns=colunas_vazias)
    validos = df[df["conta_pro_nivel"]]
    agg = validos.groupby(["medico", "anomes"]).agg(
        n_plantoes=("valor", "count"),
        n_fds=("eh_fds", "sum"),
        n_noturno=("eh_noturno", "sum"),
        n_fds_ou_noturno=("eh_fds_ou_noturno", "sum"),
        valor_repasse=("valor", "sum"),
    ).reset_index()
    coord = df[df["eh_gestao"]].groupby(["medico", "anomes"]).size().reset_index(name="_n_coord")
    agg = agg.merge(coord, on=["medico", "anomes"], how="outer")
    # medicos que so tem linha de coordenacao (sem plantao clinico contavel no mes) tambem entram
    for col in ("n_plantoes", "n_fds", "n_noturno", "n_fds_ou_noturno"):
        agg[col] = agg[col].fillna(0).astype(int)
    agg["valor_repasse"] = agg["valor_repasse"].fillna(0.0)
    agg["teve_coordenacao"] = agg["_n_coord"].fillna(0) > 0
    agg = agg.drop(columns=["_n_coord"])
    return agg


def listar_medicos(agg):
    """Lista de medicos distintos na base agregada, ordenada - usada pelo seletor da tela de
    Gestores (marcar manualmente quem entra direto no Nivel 4, sem depender do pagamento de
    Coordenacao/Gestao ter sido lancado certo na base)."""
    if agg.empty:
        return []
    return sorted(agg["medico"].unique())


MESES_TOLERANCIA_QUEDA_BENEFICIOS = 1


def calcular_niveis(agg, niveis=None, medicos_gestores=None, custo_seguro_mes=None):
    """Para cada medico, percorre cronologicamente os meses (desde o primeiro mes dele na base
    ate o mes mais recente) e calcula nivel_bruto, nivel_vestido (com carencia aplicada) e os
    custos do mes. Meses sem nenhuma linha pro medico dentro da janela viram 0 plantoes (nao
    "sai" do historico, mas conta como mes fraco pra carencia dos BENEFICIOS - ver "colchao"
    abaixo).

    "Colchao" contra queda pontual (pedido do usuario, 2026-08-20): um UNICO mes abaixo do
    volume minimo (ex.: ferias, licenca, imprevisto) NAO reseta a carencia acumulada dos
    BENEFICIOS (streaks[nivel_check]/nivel_vestido) - so reseta se acontecer por
    MESES_TOLERANCIA_QUEDA_BENEFICIOS+1 meses SEGUIDOS (hoje = 2). No mes de "colchao", o streak
    fica congelado (nao avanca, mas tambem nao volta a zero) - se o volume voltar no mes
    seguinte, continua de onde parou, como se aquele mes fraco isolado nao tivesse acontecido pro
    calculo de carencia. NAO afeta nivel_bruto/pagamento - esse continua 100% real-time, sem
    colchao nenhum (o medico so RECEBE pelo nivel que o volume daquele mes realmente sustenta;
    o colchao e so pra nao perder o acesso aos BENEFICIOS extras por um deslize pontual).

    IMPORTANTE (corrigido 2026-08-20, esclarecido pelo usuario): 'nivel_bruto' e 'nivel_vestido'
    tem papeis DIFERENTES agora - nivel_bruto (puro volume/FDS do mes, sem carencia) e quem
    determina o PAGAMENTO (custo_aumento_pct_mes/pct_aumento_exibido - o aumento no valor do
    plantao vale na hora, nao espera carencia). nivel_vestido (com carencia aplicada) continua
    so pra determinar os BENEFICIOS extras (custo_seguro_mes/tem_seguro, e a lista em
    BENEFICIOS_NIVEL usada no relatorio/PDF) - esses sim esperam a carencia. Antes dessa correcao
    os dois usavam nivel_vestido, o que segurava indevidamente tambem o pagamento.

    'niveis' permite recalcular com parametros customizados (editados na tela de Configuracoes)
    em vez dos parametros padrao aprovados (NIVEIS) - mesmo formato de lista de dicts.

    'medicos_gestores' e a lista curada manualmente na tela "Gestores" (nomes de medico) - quem
    esta nela tem Nivel 4 automatico em TODO o historico, junto com (nao substitui, conforme
    decisao do usuario em 2026-08-20) a deteccao automatica via pagamento de Coordenacao/Gestao
    (teve_coordenacao). Existe pra cobrir gestor que nunca teve um pagamento desse tipo lancado
    certo na base.

    'custo_seguro_mes' permite recalcular com o custo do pacote de seguro customizado (editado na
    tela de Regras do Programa) em vez do CUSTO_SEGURO_TOTAL_MES padrao aprovado."""
    niveis = sorted(niveis or NIVEIS, key=lambda n: n["idx"])
    nivel_por_idx = niveis_para_dict(niveis)
    medicos_gestores = set(medicos_gestores or [])
    custo_seguro_mes = CUSTO_SEGURO_TOTAL_MES if custo_seguro_mes is None else custo_seguro_mes
    if agg.empty:
        return pd.DataFrame()
    meses_todos = sorted(agg["anomes"].unique())
    mes_para_indice = {m: i for i, m in enumerate(meses_todos)}

    resultados = []
    for medico, grp in agg.groupby("medico"):
        eh_gestor_manual = medico in medicos_gestores
        grp = grp.set_index("anomes")
        primeiro_idx = min(mes_para_indice[m] for m in grp.index)
        ultimo_idx = mes_para_indice[meses_todos[-1]]  # roda ate o fim da base pra todo mundo
        streaks = {2: 0, 3: 0, 4: 0}
        meses_abaixo_seguidos = {2: 0, 3: 0, 4: 0}  # "colchao" - ver docstring da funcao
        for i in range(primeiro_idx, ultimo_idx + 1):
            am = meses_todos[i]
            if am in grp.index:
                row = grp.loc[am]
                n_plantoes = int(row["n_plantoes"])
                n_fds, n_noturno = int(row["n_fds"]), int(row["n_noturno"])
                n_fds_ou_noturno = int(row["n_fds_ou_noturno"])
                teve_coord = bool(row["teve_coordenacao"])
                valor_repasse = float(row["valor_repasse"])
            else:
                n_plantoes, n_fds, n_noturno, n_fds_ou_noturno = 0, 0, 0, 0
                teve_coord, valor_repasse = False, 0.0

            # exigencia de nivel (min_fds) conta sobre a UNIAO fds+noturno, nao so fds puro.
            # Nivel 4 automatico se teve pagamento de coordenacao/gestao naquele mes OU se esta
            # na lista manual de gestores (as duas coexistem - OR, decisao do usuario 2026-08-20).
            nivel_bruto = (
                niveis[-1]["idx"] if (teve_coord or eh_gestor_manual)
                else _nivel_bruto(n_plantoes, n_fds_ou_noturno, niveis)
            )

            for nivel_check in (2, 3, 4):
                if nivel_bruto >= nivel_check:
                    streaks[nivel_check] += 1
                    meses_abaixo_seguidos[nivel_check] = 0
                else:
                    meses_abaixo_seguidos[nivel_check] += 1
                    if meses_abaixo_seguidos[nivel_check] > MESES_TOLERANCIA_QUEDA_BENEFICIOS:
                        streaks[nivel_check] = 0
                    # dentro da tolerancia (1º mes abaixo, por padrao): streak fica CONGELADO,
                    # nem avanca nem reseta - "colchao" contra deslize pontual.

            nivel_vestido = 1
            for nivel_check in (2, 3, 4):
                carencia = nivel_por_idx[nivel_check]["carencia_meses"]
                if streaks[nivel_check] >= carencia + 1:
                    nivel_vestido = nivel_check

            # info_bruto rege o PAGAMENTO (imediato, por volume) - info_vestido rege so os
            # BENEFICIOS extras (esperam carencia). Ver docstring da funcao.
            info_bruto = nivel_por_idx[nivel_bruto]
            info_vestido = nivel_por_idx[nivel_vestido]
            custo_seguro = custo_seguro_mes if info_vestido["tem_seguro"] else 0.0
            custo_aumento_pct = valor_repasse * info_bruto["pct_aumento"]
            valor_total_geral = valor_repasse + custo_aumento_pct

            resultados.append({
                "medico": medico, "anomes": am,
                "n_plantoes": n_plantoes, "n_fds": n_fds, "n_noturno": n_noturno,
                "n_fds_ou_noturno": n_fds_ou_noturno, "teve_coordenacao": teve_coord,
                "valor_repasse": valor_repasse,
                "nivel_bruto": nivel_bruto, "nivel_vestido": nivel_vestido,
                "pct_aumento_exibido": info_bruto["pct_exibido"],
                # streak_nivel_bruto = meses consecutivos que o volume sustenta PELO MENOS o
                # nivel_bruto atual (None no Nivel 1, que e o piso e nao tem streak) - usado como
                # "tempo no nivel atual" de PAGAMENTO (sem desconto de carencia, diferente de
                # tempo_no_nivel_atual() abaixo, que e sobre o nivel_vestido/beneficios).
                "streak_nivel_bruto": streaks[nivel_bruto] if nivel_bruto >= 2 else None,
                "streak_n2": streaks[2], "streak_n3": streaks[3], "streak_n4": streaks[4],
                # meses_abaixo_n{X} = quantos meses SEGUIDOS o volume já ficou abaixo do nível X,
                # dentro da janela de tolerância do colchão (nunca passa de
                # MESES_TOLERANCIA_QUEDA_BENEFICIOS - se passasse, o streak já teria resetado e
                # nivel_vestido já teria caído). >=1 pro nivel_vestido atual = "está no mês de
                # tolerância agora, um mês fraco a mais perde o benefício" - ver "risco de queda de
                # benefício" na página Abordagem (pedido do usuário 2026-08-21).
                "meses_abaixo_n2": meses_abaixo_seguidos[2],
                "meses_abaixo_n3": meses_abaixo_seguidos[3],
                "meses_abaixo_n4": meses_abaixo_seguidos[4],
                "custo_seguro_mes": custo_seguro, "custo_aumento_pct_mes": custo_aumento_pct,
                "valor_total_geral": valor_total_geral,
            })
    return pd.DataFrame(resultados)


def montar_agregado(arquivo=None):
    """Le + agrega (a parte lenta: leitura do Excel). Nao depende dos parametros de nivel, entao
    fica cacheavel separado - so precisa rodar de novo se o Excel mudar, nao se as regras
    mudarem na tela de Configuracoes."""
    df_linhas = carregar_plantoes(arquivo)
    return agregar_mensal(df_linhas)


def montar_base_completa(arquivo=None, niveis=None, medicos_gestores=None, custo_seguro_mes=None):
    """Pipeline completo: le, agrega, calcula niveis. Cacheable pela camada de UI."""
    agg = montar_agregado(arquivo)
    return calcular_niveis(
        agg, niveis=niveis, medicos_gestores=medicos_gestores, custo_seguro_mes=custo_seguro_mes
    )


def status_atual(niveis_df, anomes_referencia=None):
    """Snapshot do mes mais recente disponivel (ou anomes_referencia se passado) por medico."""
    if niveis_df.empty:
        return niveis_df
    am = anomes_referencia or niveis_df["anomes"].max()
    return niveis_df[niveis_df["anomes"] == am].copy()


def proximo_nivel_info(row):
    """Dado um snapshot-row de status_atual, calcula quanto falta pro proximo nivel (por
    volume - carencia adicional continua sendo tempo, nao da pra "comprar"). Usa n_fds_ou_noturno
    (a uniao, mesmo campo que _nivel_bruto usa de verdade) - corrigido em 2026-08-20, a versao
    anterior comparava com n_fds puro e ficava incoerente com a regra vigente. Sem uso em nenhuma
    tela ate essa correcao (funcao morta), entao nao ha call-site desatualizado pra ajustar."""
    nivel_atual = row["nivel_bruto"]
    if nivel_atual >= 4:
        return None
    prox = NIVEL_POR_IDX[nivel_atual + 1]
    faltam_plantoes = max(0, prox["min_plantoes"] - row["n_plantoes"])
    faltam_fds_ou_noturno = max(0, prox["min_fds"] - row["n_fds_ou_noturno"])
    return {
        "proximo_nivel": prox["nome"], "faltam_plantoes": faltam_plantoes,
        "faltam_fds_ou_noturno": faltam_fds_ou_noturno,
    }


def simular_todos_niveis(row, niveis=None):
    """Generaliza proximo_nivel_info: em vez de só o próximo nível, calcula pra CADA nível ainda
    não alcançado (acima do nivel_bruto do mês) quantos plantões/FDS+Noturno faltam e uma
    estimativa de ganho extra de bonificação (R$) se o médico completasse esse volume ainda este
    mês. Pedido do usuário (2026-08-20): dar ao escalista uma visão tipo "faltam X plantões pro N3,
    mais Y pro N4" pra usar no dia a dia, junto com o valor estimado.

    O ganho estimado usa o mesmo raciocínio de proximo_nivel_info/renderizar_relatorio_medico: o
    ticket médio já realizado no mês (valor_repasse / n_plantoes) projeta o valor com os plantões
    que faltam, aplica o % daquele nível, e subtrai a bonificação que já está valendo agora
    (custo_aumento_pct_mes, calculada em cima do nivel_vestido, não do nivel_bruto - pode haver
    carência entre os dois). É estimativa (o ticket médio dos próximos plantões pode variar), não
    projeção garantida.

    Retorna lista ordenada do nível mais próximo pro mais distante; lista vazia se já está no
    Nível 4."""
    niveis = niveis or NIVEIS
    niveis_dict = {n["idx"]: n for n in niveis}
    nivel_atual = int(row["nivel_bruto"])
    n_plantoes_atual = int(row["n_plantoes"])
    valor_repasse_atual = float(row["valor_repasse"])
    ticket_medio = valor_repasse_atual / n_plantoes_atual if n_plantoes_atual > 0 else 0.0
    bonificacao_atual = float(row.get("custo_aumento_pct_mes", 0.0))

    resultado = []
    for idx in range(nivel_atual + 1, 5):
        alvo = niveis_dict.get(idx)
        if alvo is None:
            continue
        faltam_plantoes = max(0, alvo["min_plantoes"] - n_plantoes_atual)
        faltam_fds_ou_noturno = max(0, alvo["min_fds"] - row["n_fds_ou_noturno"])
        valor_projetado = valor_repasse_atual + faltam_plantoes * ticket_medio
        ganho_projetado = valor_projetado * alvo["pct_aumento"]
        resultado.append({
            "nivel_idx": idx, "nivel_nome": alvo["nome"], "pct_exibido": alvo["pct_exibido"],
            "faltam_plantoes": faltam_plantoes, "faltam_fds_ou_noturno": faltam_fds_ou_noturno,
            "ganho_extra_estimado": max(0.0, ganho_projetado - bonificacao_atual),
        })
    return resultado


# Beneficios por nivel (Programa_Constelacao_CallMed_v2.md secao 2) - cada nivel soma aos
# anteriores (Nivel 3 tem tudo do 1+2+3). Usado no "Relatorio do Medico" (tela + PDF pra entregar
# ao medico) - texto em tom de comunicado, nao o tom tecnico do documento de regras.
BENEFICIOS_NIVEL = {
    1: [
        "Suporte administrativo prioritário",
        "Apoio jurídico integral (retroatividade de 3 anos para fatos desconhecidos)",
        "Desconto no convênio médico ao contratar pela CallMed",
        "Antecipação de valores de plantão (pago até 2 dias após o plantão)",
    ],
    # Pacote de seguros movido do Nivel 3 pro Nivel 2 (mudanca de regra pedida pelo usuario,
    # 2026-08-20) - ver tem_seguro em NIVEIS acima.
    2: [
        "Desconto em planos de saúde",
        "Antecipação dos plantões agendados",
        "Acesso ao programa de mindfulness",
        "Seguro de Vida — R$ 99.000 (morte natural) / R$ 198.000 (morte acidental)",
        "DIT — R$ 166,66/dia em caso de incapacidade temporária",
        "Assistência Funeral Familiar — R$ 10.000",
        "RCP — Responsabilidade Civil Profissional, R$ 100.000, franquia zero",
    ],
    3: [
        "50% de reembolso em cursos (ACLS/PALS/COPA/SAVA, 1x/ano, até 3 localidades indicadas)",
        "Escala preferencial em unidades",
    ],
    4: [
        "100% de reembolso em cursos",
        "15 dias de licença maternidade remunerada",
        "7 dias de licença paternidade remunerada",
        "Reconhecimento oficial em eventos CallMed",
    ],
}


def beneficios_acumulados(nivel):
    """Lista cumulativa de beneficios do Nivel 1 ate o nivel informado (cada nivel inclui tudo
    dos anteriores + os proprios)."""
    beneficios = []
    for n in range(1, int(nivel) + 1):
        beneficios.extend(BENEFICIOS_NIVEL.get(n, []))
    return beneficios


def tempo_no_nivel_atual(row, niveis=None):
    """Quantos meses o BENEFICIO (seguro, curso, licenca - nao o pagamento, ver calcular_niveis)
    do nivel_vestido esta realmente ativo (streak do nivel menos a carencia que ja foi cumprida
    pra ele valer) - diferente do streak bruto, que conta tambem os meses "gastos" na propria
    carencia. Nivel 1 nao tem streak/carencia (e o piso), retorna None. Pro "tempo no nivel
    atual" de PAGAMENTO (sem desconto de carencia), usar o campo streak_nivel_bruto direto."""
    niveis = niveis or NIVEIS
    nivel_vestido = int(row["nivel_vestido"])
    if nivel_vestido < 2:
        return None
    carencia = niveis_para_dict(niveis)[nivel_vestido]["carencia_meses"]
    streak = int(row.get(f"streak_n{nivel_vestido}", 0))
    return max(1, streak - carencia)


def risco_queda_beneficio(row, niveis=None):
    """Dado um snapshot-row de status_atual, avalia se o médico está DENTRO do mês de tolerância
    do colchão pro nivel_vestido atual - ou seja, o volume deste mês já ficou abaixo do mínimo do
    nível de benefício, mas o "colchão" ainda está segurando (não resetou a carência ainda). Um
    mês fraco A MAIS reseta o streak e derruba o benefício pra valer (ver
    MESES_TOLERANCIA_QUEDA_BENEFICIOS/calcular_niveis). Retorna None se não estiver em risco
    (Nível 1, sem carência pra perder, ou o mês atual já sustentou o volume exigido).

    Pedido do usuário (2026-08-21): contraparte da "Abordagem (quase lá)" - dar visibilidade de
    quem está prestes a CAIR de nível, não só quem está prestes a SUBIR, pro escalista poder agir
    antes de perder o benefício (não antes de ganhar)."""
    niveis = niveis or NIVEIS
    nivel_vestido = int(row["nivel_vestido"])
    if nivel_vestido < 2:
        return None
    meses_abaixo = int(row.get(f"meses_abaixo_n{nivel_vestido}", 0))
    if meses_abaixo < 1:
        return None
    alvo = niveis_para_dict(niveis)[nivel_vestido]
    faltam_plantoes = max(0, alvo["min_plantoes"] - int(row["n_plantoes"]))
    faltam_fds_ou_noturno = max(0, alvo["min_fds"] - int(row["n_fds_ou_noturno"]))
    return {
        "nivel_beneficio": nivel_vestido,
        "nivel_bruto_atual": int(row["nivel_bruto"]),
        "meses_tolerancia_consumidos": meses_abaixo,
        "meses_tolerancia_total": MESES_TOLERANCIA_QUEDA_BENEFICIOS,
        "faltam_plantoes": faltam_plantoes,
        "faltam_fds_ou_noturno": faltam_fds_ou_noturno,
    }


def custo_por_operacao_mes(df_linhas, niveis_df, rampup_df, mes_ref):
    """Quebra o custo do programa (aumento % + ramp-up) por operação/hospital no mês de
    referência - pedido do usuário (2026-08-21), pra enxergar onde o programa concentra custo, não
    só o total agregado.

    O bônus de ramp-up já É por operação por construção (o disparo escolhe a operação) - soma
    direta, exata. Já o % de aumento é um atributo do MÉDICO/nível (calculado em cima do repasse
    TOTAL do médico no mês, todas as operações somadas) - não existe uma "fatia exata" dele por
    hospital quando o médico atende mais de um. Aqui ele é ALOCADO proporcionalmente ao share de
    valor_repasse de cada operação dentro do repasse elegível do médico naquele mês - é uma
    estimativa de concentração, não uma contabilização exata por hospital (mesmo espírito do "%
    aproximado" já sinalizado em outras telas)."""
    colunas_vazias = ["operacao", "custo_aumento_alocado", "custo_rampup", "custo_total", "n_medicos"]
    if df_linhas.empty or niveis_df.empty:
        return pd.DataFrame(columns=colunas_vazias)

    linhas_mes = df_linhas[(df_linhas["anomes"] == mes_ref) & df_linhas["conta_pro_nivel"]]
    if linhas_mes.empty:
        return pd.DataFrame(columns=colunas_vazias)

    # Share de cada operação no repasse do médico naquele mês (base do rateio do aumento %).
    repasse_por_op = linhas_mes.groupby(["medico", "operacao"])["valor"].sum().reset_index()
    repasse_total_medico = repasse_por_op.groupby("medico")["valor"].transform("sum")
    repasse_por_op["share"] = repasse_por_op["valor"] / repasse_total_medico

    custo_medico_mes = niveis_df.loc[niveis_df["anomes"] == mes_ref, ["medico", "custo_aumento_pct_mes"]]
    repasse_por_op = repasse_por_op.merge(custo_medico_mes, on="medico", how="left")
    repasse_por_op["custo_aumento_pct_mes"] = repasse_por_op["custo_aumento_pct_mes"].fillna(0.0)
    repasse_por_op["custo_aumento_alocado"] = repasse_por_op["custo_aumento_pct_mes"] * repasse_por_op["share"]

    aumento_por_op = repasse_por_op.groupby("operacao").agg(
        custo_aumento_alocado=("custo_aumento_alocado", "sum"),
        n_medicos=("medico", "nunique"),
    ).reset_index()

    # Ramp-up: exato, sem rateio - a operação já vem do próprio disparo.
    partes_rampup = []
    if rampup_df is not None and not rampup_df.empty:
        for _, disparo in rampup_df.iterrows():
            meses_disparo = set(disparo["meses"] or [])
            if mes_ref not in meses_disparo:
                continue
            valor_op = linhas_mes.loc[linhas_mes["operacao"] == disparo["operacao"], "valor"].sum()
            if valor_op:
                partes_rampup.append({"operacao": disparo["operacao"], "custo_rampup": valor_op * float(disparo["pct"])})
    rampup_por_op = (
        pd.DataFrame(partes_rampup).groupby("operacao", as_index=False)["custo_rampup"].sum()
        if partes_rampup else pd.DataFrame(columns=["operacao", "custo_rampup"])
    )

    resultado = aumento_por_op.merge(rampup_por_op, on="operacao", how="outer")
    resultado["custo_aumento_alocado"] = resultado["custo_aumento_alocado"].fillna(0.0)
    resultado["custo_rampup"] = resultado["custo_rampup"].fillna(0.0)
    resultado["n_medicos"] = resultado["n_medicos"].fillna(0).astype(int)
    resultado["custo_total"] = resultado["custo_aumento_alocado"] + resultado["custo_rampup"]
    return resultado.sort_values("custo_total", ascending=False).reset_index(drop=True)


def meses_por_nivel(hist_medico, campo="nivel_bruto"):
    """Conta quantos meses (nao necessariamente consecutivos) o medico passou em cada nivel ao
    longo de todo o historico disponivel (hist_medico = niveis_df filtrado por 1 medico, jah
    ordenado ou nao - nao importa aqui). Usa nivel_bruto por padrao (nivel de PAGAMENTO, sem
    carencia) - pedido do usuario 2026-08-20, pra mostrar no relatorio "quantos meses ele ficou
    em cada nivel". Passe campo='nivel_vestido' se precisar da contagem por nivel de BENEFICIOS."""
    contagem = hist_medico[campo].value_counts().to_dict()
    return {n: int(contagem.get(n, 0)) for n in (1, 2, 3, 4)}
